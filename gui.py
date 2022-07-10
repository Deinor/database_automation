from operator import index
import PyQt5.QtWidgets as qtw
import PyQt5.QtGui as qtg
import openpyxl
import datetime
import calendar
from openpyxl.utils.cell import coordinate_from_string
from excel_script import get_dataFC, used_dataFC

#Load curent date, curent month andd previous month
current_date = datetime.datetime.now()      #Curent date.time type of return
curent_month = current_date.strftime("%B")  #Curent month date.time type of return 
previous_month = calendar.month_name[datetime.datetime.now().month - 1] #Previous month
if previous_month == 0:
    previous_month =12

#Load data from final-confirmed excel
wb_FC = openpyxl.load_workbook('final-confirmed_2022_6_23.xlsx')
wb_FC.sheetnames           # The workbook's sheets' names.
sheet_FC = wb_FC ["CAD"]      #Get a sheet from workbook  
#print(sheet_FC)            #Debug_1

#Load data from yellow-blue excel
#wb_YB = openpyxl.load_workbook('yellow-blue_2022_6_23.xlsx')
#wb_YB.sheetnames           # The workbook's sheets' names.
#sheet_YB = wb_YB ["CAD"]      #Get a sheet from workbook  
#print(sheet_YB)            #Debug_1
 
# Get list of projects
s_nr_col = sheet_FC ["C"]
list_snr = []
for x in range(4, len(s_nr_col)):       #S-number list
    list_snr += [s_nr_col[x].value]

#print(list_snr)                         #Debug_1_1

#Load empty used_data dictionary
dic_FC = used_dataFC
    
class MainWindow(qtw.QWidget):
    def __init__(self):
        super().__init__()
        
        #Add Title
        self.setWindowTitle("Johny XXX")

        #Midnight Pebble Beach
        color_window =  """
            QWidget {
                background-color: rgb(237,245,244);
                }
            """
        color_button = """
            QWidget {
                border-radius: 8px;
                background-color: rgb(46,92,87);
                font-size: 18px;
                padding: 10px;
                font-weight: 500;
                color: white
                }
            QPushButton:hover {
                border: 4px solid rgb(141,205,198);
                border-radius: 8px;
                background-color: rgb(93,184,174);
                font-size: 18px;
                color: black               
                }
            """
        color_table = """
            QWidget {
                background-color: rgb(56,112,106);
                color: white;
                font-weight: 500;
                }
            QHeaderView:section{
                Background-color:rgb(7,17,37);
                font-weight: 500;
                border-radius: 15px;
            }
        """
        color_label = """
            QWidget {
                background-color: rgb(133,189,182);
                font-size: 22px;
                color: black;
                border-radius: 15px;
                padding: 5px;
                font-weight: 500;
                }
        """
        color_combobox = """
              QWidget {
                background-color: rgb(133,189,182);
                padding: 5px;
                padding-left: 20px;
                color: black;
                border-radius: 15px;
              }  
        """

        #Creating layouts
        self.setGeometry(600, 300, 800, 600)
        outer_layout = qtw.QVBoxLayout()
        top_layout = qtw.QFormLayout()
        body_layout = qtw.QVBoxLayout()
        self.setStyleSheet(color_window)

        #Nesting layouts
        outer_layout.addLayout(top_layout)
        outer_layout.addLayout(body_layout)

        #Set the window´s main layout
        self.setLayout(outer_layout)

        #Create lable for list SNR
        label_snr = qtw.QLabel("Select S-Number:")
        label_snr.setStyleSheet(color_label)
       
        #Change font, text size of lable
        label_snr.setFont(qtg.QFont("Helvetica", 15))

        top_layout.addWidget(label_snr)

        #Create Combo box
        my_combo = qtw.QComboBox(self)

        #Add Items to combo Box
        my_combo.addItems(list_snr)
        my_combo.setStyleSheet(color_combobox)

        #Change font, text size of lable
        my_combo.setFont(qtg.QFont("Helvetica", 10))

        #Put combo on screen
        top_layout.addWidget(my_combo)

        #Conect signals to the metods - get index of list
        def comboBox_activated(index):
            global list_snrIndex
            list_snrIndex = index
            generate_data()
            
        my_combo.activated.connect(comboBox_activated)

        #Create table 
        table1 = qtw.QTableWidget(self)
        table1.setRowCount(1)
        table1.setColumnCount(7)
        table1.setHorizontalHeaderLabels(["Project number", "S-Number", "Project leader DE", "Ordered hours", "Used hours", "Available hours", "Used hours in last month"])
        table1.verticalHeader().setVisible(False)
        table1.setStyleSheet(color_table)
        table1.horizontalHeader().setStyleSheet(color_table)

        #Fill table with dictionary values andd keys 
        table1.setItem(0, 0, qtw.QTableWidgetItem(str(dic_FC["project_number"])))
        table1.setItem(0, 1, qtw.QTableWidgetItem(str(dic_FC["s_nr"])))
        table1.setItem(0, 2, qtw.QTableWidgetItem(str(dic_FC["pjm_de"])))
        table1.setItem(0, 3, qtw.QTableWidgetItem(str(dic_FC["ordered"])))
        table1.setItem(0, 4, qtw.QTableWidgetItem(str(dic_FC["used_hours"])))
        table1.setItem(0, 5, qtw.QTableWidgetItem(str(dic_FC["available"])))
        table1.setItem(0, 6, qtw.QTableWidgetItem(str(dic_FC["used_lastm_hours"])))
            
        
        #Put table on screen
        body_layout.addWidget(table1)
        
        #Create a button
        button_generateOut = qtw.QPushButton("Generate e-mail", clicked = lambda: generate_email())
         
        button_generateOut.setStyleSheet(color_button)
        body_layout.addWidget(button_generateOut)

        #Generate out. data
        def generate_data():

            cell_address = "C" + str(list_snrIndex + 5)
            #print(cell_address) #Debug 2_2

            #Get selected row_FC number

            xy = coordinate_from_string (cell_address)
            row_FC = xy [1]
            #print(row_FC)          #Debug 2_3
        
            dic_FC = get_dataFC(row_FC, sheet_FC, curent_month)
            #print(dic_FC) #Debug 2_4

            #Update table dictionary values each time project is selected in combo box
            table1.setItem(0, 0, qtw.QTableWidgetItem(str(dic_FC["project_number"])))
            table1.setItem(0, 1, qtw.QTableWidgetItem(str(dic_FC["used_hours"])))
            table1.setItem(0, 1, qtw.QTableWidgetItem(str(dic_FC["s_nr"])))
            table1.setItem(0, 2, qtw.QTableWidgetItem(str(dic_FC["pjm_de"])))
            table1.setItem(0, 3, qtw.QTableWidgetItem(str(dic_FC["ordered"])))
            table1.setItem(0, 4, qtw.QTableWidgetItem(str(dic_FC["used_hours"])))
            table1.setItem(0, 5, qtw.QTableWidgetItem(str(dic_FC["available"])))
            table1.setItem(0, 6, qtw.QTableWidgetItem(str(dic_FC["used_lastm_hours"])))
            

         #Generate email
        def generate_email():
            print("E-mail generated!", used_dataFC)

        #Show the all
        self.show()

app = qtw.QApplication([])
mw = MainWindow()

#Run window
app.exec_()